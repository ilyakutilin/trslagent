import csv
import os
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

from src.config import get_settings, logger

settings = get_settings()


class GlossaryParser:
    def __init__(self, file_path: str) -> None:
        self.file_path = file_path

    def _load_terms_from_csv(self, csv_path: Path) -> list[dict]:
        """Read CSV into terms dict list."""
        terms = []
        stem = csv_path.stem
        with open(csv_path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            # Skip header row
            next(reader, None)
            for row in reader:
                if len(row) < 3:
                    logger.warning(f"Row has less than 3 columns: {' | '.join(row)}")
                    continue
                try:
                    term_id = int(row[0].strip())
                except (ValueError, TypeError):
                    raise ValueError(
                        f"Invalid ID value '{row[0].strip()}' in CSV. "
                        "ID must be a valid integer."
                    )
                source, target = row[1].strip(), row[2].strip()
                if source and target:
                    terms.append(
                        {"id": f"{stem}_{term_id}", "source": source, "target": target}
                    )
        logger.info(f"Read {len(terms)} terms from CSV.")
        return terms

    def _load_terms_from_xlsx(self, xlsx_path: Path) -> list[dict]:
        """Read XLSX into terms dict list."""
        terms = []

        if not os.path.exists(xlsx_path):
            raise FileNotFoundError(f"XLSX file not found: {xlsx_path}")

        stem = xlsx_path.stem

        try:
            workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
            sheet = workbook.active

            if sheet is None:
                workbook.close()
                raise ValueError("XLSX file is empty or has no active sheet.")

            # Skip header row by enumerating and skipping index 0
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:  # Skip header row
                    continue
                if not row or len(row) < 3:
                    continue
                try:
                    term_id = int(str(row[0]).strip())
                except (ValueError, TypeError):
                    raise ValueError(
                        f"Invalid ID value '{row[0]}' in XLSX at row {i + 1}. "
                        "ID must be a valid integer."
                    )
                source, target = str(row[1]).strip(), str(row[2]).strip()
                if source and target:
                    terms.append(
                        {"id": f"{stem}_{term_id}", "source": source, "target": target}
                    )

        finally:
            if "workbook" in locals():
                workbook.close()

        logger.info(f"Read {len(terms)} terms from XLSX.")
        return terms

    def _load_terms_from_xml(self, xml_path: Path) -> list[dict]:
        """
        Read a MultiTerm XML file into terms dict list.

        Each conceptGrp in the XML becomes one entry in the terms dict list with:
        - "id":     "<filename_stem>_<concept_id>"
        - "source": all terms in the first language joined by " | "
        - "target": all terms in the second language joined by " | "

        Args:
            xml_path: Absolute or relative path to the .xml glossary file.

        Raises:
            FileNotFoundError: If the file does not exist.
            ValueError:        If the XML structure is invalid or inconsistent.
            RuntimeError:      For unexpected errors during parsing.
        """
        # ------------------------------------------------------------------ #
        # 1. Resolve path and derive the file-name stem used as ID prefix     #
        # ------------------------------------------------------------------ #
        file_stem = xml_path.stem
        logger.info(f"Loading glossary terms from XML: {xml_path} (stem={file_stem})")

        if not xml_path.is_file():
            logger.error(f"XML file not found: {xml_path}")
            raise FileNotFoundError(f"Glossary XML file not found: {xml_path}")

        # ------------------------------------------------------------------ #
        # 2. Parse the XML document                                           #
        # ------------------------------------------------------------------ #
        try:
            tree = ET.parse(xml_path)
        except ET.ParseError as exc:
            logger.error(f"Malformed XML in {xml_path}: {exc}")
            raise ValueError(f"Failed to parse XML file {xml_path}: {exc}") from exc
        except OSError as exc:
            logger.error(f"Cannot read file {xml_path}: {exc}")
            raise RuntimeError(f"Cannot read glossary file {xml_path}: {exc}") from exc

        root = tree.getroot()
        if root.tag != "mtf":
            logger.error(
                f"Unexpected root tag {root.tag} in {xml_path} (expected 'mtf')"
            )
            raise ValueError(
                f"Invalid glossary XML: root tag is {root.tag}, expected 'mtf'."
            )

        # ------------------------------------------------------------------ #
        # 3. Detect the canonical source / target language pair from the      #
        #    first conceptGrp that has at least two distinct language codes.  #
        # ------------------------------------------------------------------ #
        source_lang: str | None = None  # e.g. "EN"
        target_lang: str | None = None  # e.g. "RU"

        concept_groups = root.findall("conceptGrp")
        if not concept_groups:
            logger.warning(
                f"No <conceptGrp> elements found in {xml_path}; list is empty."
            )
            return []

        for cg in concept_groups:
            langs_seen: list[str] = []
            for lg in cg.findall("languageGrp"):
                lang_el = lg.find("language")
                if lang_el is None:
                    continue
                code = lang_el.get("lang", "").strip().upper()
                if code and code not in langs_seen:
                    langs_seen.append(code)
                if len(langs_seen) >= 2:
                    break
            if len(langs_seen) >= 2:
                source_lang, target_lang = langs_seen[0], langs_seen[1]
                logger.debug(
                    "Detected language pair from first qualifying conceptGrp: "
                    f"source={source_lang}, target={target_lang}"
                )
                break

        if source_lang is None or target_lang is None:
            logger.error(
                f"Could not detect a source/target language pair in {xml_path}"
            )
            raise ValueError(
                f"Glossary XML {xml_path} does not contain at least two distinct "
                "language codes in any <conceptGrp>."
            )

        # ------------------------------------------------------------------ #
        # 4. Iterate over every <conceptGrp> and build term dicts             #
        # ------------------------------------------------------------------ #
        loaded: list[dict] = []

        for idx, concept_grp in enumerate(concept_groups):
            # --- 4a. Extract the concept ID -------------------------------- #
            concept_el = concept_grp.find("concept")
            if concept_el is None or not (concept_el.text or "").strip():
                logger.warning(
                    f"conceptGrp #{idx} in {xml_path} has no <concept> id — skipping."
                )
                continue

            assert concept_el.text is not None
            raw_id = concept_el.text.strip()
            entry_id = f"{file_stem}_{raw_id}"

            # --- 4b. Collect terms per language code ----------------------- #
            terms_by_lang: dict[str, list[str]] = {}

            for lg in concept_grp.findall("languageGrp"):
                lang_el = lg.find("language")
                if lang_el is None:
                    logger.debug(
                        f"conceptGrp {entry_id}: <languageGrp> missing <language> tag "
                        "— skipped."
                    )
                    continue

                code = lang_el.get("lang", "").strip().upper()
                if not code:
                    logger.debug(
                        f"conceptGrp {entry_id}: <language> element "
                        "has no 'lang' attribute — skipped."
                    )
                    continue

                # Warn about unexpected language codes but keep the data
                if code not in (source_lang, target_lang):
                    logger.warning(
                        f"conceptGrp {entry_id}: unexpected language code {code} "
                        f"(expected {source_lang} or {target_lang}) — ignoring."
                    )
                    continue

                term_grp = lg.findall("termGrp")
                if not term_grp:
                    logger.debug(
                        f"conceptGrp {entry_id}, lang {code}: "
                        "missing <termGrp> — skipped.",
                    )
                    continue

                for term in term_grp:
                    term_el = term.find("term")
                    if term_el is None or not (term_el.text or "").strip():
                        logger.debug(
                            f"conceptGrp {entry_id}, lang {code}: "
                            "empty or missing <term> — skipped."
                        )
                        continue

                    assert term_el.text is not None
                    term_text = term_el.text.strip()
                    terms_by_lang.setdefault(code, []).append(term_text)

            # --- 4c. Validate that both languages are present -------------- #
            if source_lang not in terms_by_lang:
                logger.warning(
                    f"conceptGrp {entry_id}: no terms found "
                    f"for source language {source_lang} — skipping entry."
                )
                continue

            if target_lang not in terms_by_lang:
                logger.warning(
                    f"conceptGrp {entry_id}: no terms found "
                    "for target language {target_lang} — skipping entry."
                )
                continue

            # --- 4d. Join multiple terms with the agreed delimiter --------- #
            source_text = " | ".join(terms_by_lang[source_lang])
            target_text = " | ".join(terms_by_lang[target_lang])

            entry = {"id": entry_id, "source": source_text, "target": target_text}
            loaded.append(entry)
            logger.debug(f"Parsed entry: {entry}")

        # ------------------------------------------------------------------ #
        # 5. Commit results                                                   #
        # ------------------------------------------------------------------ #
        logger.info(
            f"Finished loading {len(loaded)} term(s) from {xml_path} "
            f"(source={source_lang}, target={target_lang})."
        )
        return loaded

    def load_terms_from_file(self) -> list[dict]:
        """Auto-detect file format and load accordingly."""
        path = Path(self.file_path)
        ext = path.suffix
        if ext == ".csv":
            return self._load_terms_from_csv(path)
        elif ext == ".xlsx":
            return self._load_terms_from_xlsx(path)
        elif ext == ".xml":
            return self._load_terms_from_xml(path)
        else:
            raise ValueError(
                f"Unsupported file format '{ext}'. "
                "Only .csv, .xlsx, or .xml files are supported."
            )
