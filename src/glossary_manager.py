"""
Glossary Manager
================
Loads a glossary CSV or XLSX, embeds all terms using multilingual-e5-large,
stores them in ChromaDB, and provides RAG-style retrieval per chunk.

Glossary file format (no header required, but header is fine):
  Three columns: id, source_term, target_term
  - id: unique integer identifier for each term pair
  - source_term: term in source language
  - target_term: term in target language
  Supported formats: .csv, .xlsx
  e.g.:
  1,contract termination,Vertragsauflösung
  2,force majeure,höhere Gewalt

The glossary is bidirectional: querying in either language will find the entry.
"""

import csv
import os
from typing import Optional
from xml.etree import ElementTree as ET

import chromadb
from chromadb.config import Settings
from openpyxl import load_workbook
from sentence_transformers import SentenceTransformer

from src.config import get_settings, logger

# ── Configuration ─────────────────────────────────────────────────────────────

settings = get_settings()


# ── Why multilingual-e5-large? ────────────────────────────────────────────────
#
# - Trained on 100+ languages including all major European/Asian languages
# - 'e5' means it's an "embed everything" model optimised for retrieval tasks
# - The 'large' variant (560M params) gives better accuracy than 'small'/'base'
#   at the cost of ~2× slower embedding. For a one-time glossary embed, fine.
# - Crucially: it maps equivalent terms in different languages to nearby vectors,
#   so querying with an English chunk will still retrieve German glossary entries.
# - Important: e5 models expect a prefix on the text:
#     "query: <text>"  for things you're searching WITH
#     "passage: <text>" for things you're searching FOR (the glossary entries)
#   We handle this automatically below.


class GlossaryManager:
    def __init__(self, chroma_dir: str = settings.glossary.chroma_dir):
        self.chroma_dir = chroma_dir
        self.model: Optional[SentenceTransformer] = None
        self.collection = None
        self._terms: list[dict] = []  # in-memory copy for quick lookup by id

    # ── Lazy model loading ────────────────────────────────────────────────────

    def _get_model(self) -> SentenceTransformer:
        """Load the embedding model once, reuse thereafter."""
        if self.model is None:
            logger.info(
                f"Loading embedding model '{settings.glossary.embedding_model}' "
                "(first run downloads ~2GB)..."
            )
            self.model = SentenceTransformer(settings.glossary.embedding_model)
            logger.info(
                f"Embedding model '{settings.glossary.embedding_model}' is loaded."
            )
        return self.model

    # ── ChromaDB setup ────────────────────────────────────────────────────────

    def _get_collection(self):
        if self.collection is None:
            client = chromadb.PersistentClient(
                path=self.chroma_dir,
                settings=Settings(anonymized_telemetry=False),
            )
            self.collection = client.get_or_create_collection(
                name=settings.glossary.collection_name,
                metadata={"hnsw:space": "cosine"},  # cosine similarity for text
            )
            logger.info(f"ChromaDB collection '{self.collection.name}' is loaded.")
        return self.collection

    # ── Glossary loading & embedding ──────────────────────────────────────────

    def sync_glossary(self, glossary_source_path: str, sync_mode: bool = False) -> None:
        """
        Reads glossary CSV or XLSX and syncs terms into ChromaDB.

        When sync_mode=False (default):
            - If DB is empty, loads and embeds all terms
            - If DB has entries, just loads terms into memory (skips embedding)

        When sync_mode=True:
            - Treats the glossary source as the source of truth
            - Adds new terms (IDs not in DB)
            - Deletes removed terms (IDs in DB but not in source)
            - Updates modified terms (same ID but different source/target)
        """
        collection = self._get_collection()

        # Load terms from file first (needed for both modes)
        self._load_terms_from_file(glossary_source_path)

        if not self._terms:
            raise ValueError("Glossary is empty. Check your file.")

        # Validate IDs before proceeding
        self._validate_ids()

        # If sync not requested and DB has entries, just return after loading terms
        if not sync_mode and collection.count() > 0:
            logger.info(
                f"Glossary already loaded ({collection.count()} entries). "
                "Skipping sync step. Pass sync_mode=True to sync."
            )
            return

        # Get current state
        source_ids = {term["id"] for term in self._terms}
        db_ids = self._get_existing_db_ids()

        # Determine what needs to be added, deleted, or updated
        new_ids = source_ids - db_ids
        deleted_ids = db_ids - source_ids
        common_ids = source_ids & db_ids

        # Check for modified terms among common IDs
        modified_ids = self._get_modified_ids(common_ids)

        # Perform sync operations
        if new_ids:
            logger.info(f"Adding {len(new_ids)} new term(s) to glossary...")
            self._sync_add_terms(new_ids)

        if deleted_ids:
            logger.info(f"Deleting {len(deleted_ids)} term(s) from glossary...")
            self._sync_delete_terms(deleted_ids)

        if modified_ids:
            logger.info(f"Updating {len(modified_ids)} modified term(s)...")
            self._sync_update_terms(modified_ids)

        # If DB was empty, embed everything
        if not db_ids:
            logger.info(f"Loading glossary from {glossary_source_path}...")
            self._embed_all_terms()

        logger.info(
            f"Glossary sync complete. Total entries in DB: {collection.count()}"
        )

    def _get_existing_db_ids(self) -> set[int]:
        """Extract all term IDs currently in the ChromaDB collection."""
        collection = self._get_collection()
        result = collection.get(include=[])
        ids = (result or {}).get("ids", [])

        # Parse IDs like "fwd_123" or "rev_123" to extract integer ID
        term_ids = set()
        for entry_id in ids:
            # IDs are formatted as "fwd_{id}" or "rev_{id}"
            if entry_id.startswith("fwd_") or entry_id.startswith("rev_"):
                try:
                    term_id = int(entry_id.split("_", 1)[1])
                    term_ids.add(term_id)
                except (ValueError, IndexError):
                    continue

        return term_ids

    def _get_modified_ids(self, common_ids: set[int]) -> list[int]:
        """Find IDs where source or target terms have changed."""
        # Build lookup from DB metadata
        collection = self._get_collection()
        modified_ids = []

        # Get metadata for all fwd entries (we only need one direction to check)
        result = collection.get(where={"direction": "fwd"}, include=["metadatas"])
        all_metadata = (result or {}).get("metadatas") or []

        # Build dict of id -> (source, target)
        db_terms = {}
        for meta in all_metadata:
            term_id = meta.get("id")
            if term_id is not None:
                db_terms[term_id] = (meta.get("source"), meta.get("target"))

        # Build lookup from loaded terms
        source_terms = {
            term["id"]: (term["source"], term["target"]) for term in self._terms
        }

        # Compare
        for term_id in common_ids:
            if term_id in db_terms and term_id in source_terms:
                db_src, db_tgt = db_terms[term_id]
                src_src, src_tgt = source_terms[term_id]
                if db_src != src_src or db_tgt != src_tgt:
                    modified_ids.append(term_id)

        return modified_ids

    def _embed_and_upsert_terms(self, terms: list[dict]) -> None:
        """Embed terms and upsert into ChromaDB. Handles both forward and reverse entries.

        Processes terms in batches to avoid losing all progress if something fails.
        Each batch is embedded and upserted immediately, so partial progress is preserved.
        """
        model = self._get_model()
        collection = self._get_collection()

        total_terms = len(terms)
        batch_size = settings.glossary.embed_batch_size
        total_entries = total_terms * 2  # Each term produces 2 entries (fwd + rev)

        # Track cumulative progress
        entries_processed = 0

        # Process terms in batches
        for batch_start in range(0, total_terms, batch_size):
            batch_end = min(batch_start + batch_size, total_terms)
            batch = terms[batch_start:batch_end]

            ids, embeddings, documents, metadatas = [], [], [], []

            for term in batch:
                term_id = term["id"]
                src = term["source"]
                tgt = term["target"]
                canonical_src, canonical_tgt = sorted([src, tgt])

                # Forward entry (source → target)
                fwd_text = f"passage: {src}"
                fwd_embedding = model.encode(
                    fwd_text, normalize_embeddings=True
                ).tolist()
                ids.append(f"fwd_{term_id}")
                embeddings.append(fwd_embedding)
                documents.append(src)
                metadatas.append(
                    {
                        "source": src,
                        "target": tgt,
                        "direction": "fwd",
                        "canonical_source": canonical_src,
                        "canonical_target": canonical_tgt,
                        "id": term_id,
                    }
                )

                # Reverse entry (target → source) — same pair, but queryable from target lang
                rev_text = f"passage: {tgt}"
                rev_embedding = model.encode(
                    rev_text, normalize_embeddings=True
                ).tolist()
                ids.append(f"rev_{term_id}")
                embeddings.append(rev_embedding)
                documents.append(tgt)
                metadatas.append(
                    {
                        "source": tgt,
                        "target": src,
                        "direction": "rev",
                        "canonical_source": canonical_src,
                        "canonical_target": canonical_tgt,
                        "id": term_id,
                    }
                )

                logger.debug(
                    f"Embedding term {term_id}: {canonical_src} = {canonical_tgt}"
                )

            # Upsert this batch immediately
            try:
                collection.upsert(
                    ids=ids,
                    embeddings=embeddings,
                    documents=documents,
                    metadatas=metadatas,
                )
            except Exception as e:
                entries_so_far = entries_processed
                raise RuntimeError(
                    f"Failed to upsert batch starting at term index {batch_start}. "
                    f"Entries processed before failure: {entries_so_far}/{total_entries}"
                ) from e

            # Update cumulative progress
            entries_processed += len(ids)
            progress_pct = (entries_processed / total_entries) * 100
            logger.info(
                f"  Upserted {entries_processed}/{total_entries} entries ({progress_pct:.1f}%)..."
            )

    def _sync_add_terms(self, term_ids: set[int]) -> None:
        """Add embeddings for new term IDs."""
        # Filter terms to only those with IDs in term_ids
        new_terms = [term for term in self._terms if term["id"] in term_ids]
        self._embed_and_upsert_terms(new_terms)
        logger.info(f"  Added {len(new_terms)} new term(s) to glossary.")

    def _sync_delete_terms(self, term_ids: set[int]) -> None:
        """Delete embeddings for removed term IDs."""
        collection = self._get_collection()

        ids_to_delete = []
        for term_id in term_ids:
            ids_to_delete.append(f"fwd_{term_id}")
            ids_to_delete.append(f"rev_{term_id}")

        collection.delete(ids=ids_to_delete)
        logger.info(f"  Deleted {len(term_ids)} term(s) from glossary.")

    def _sync_update_terms(self, term_ids: list[int]) -> None:
        """Update embeddings for modified term IDs."""
        # Delete old entries
        self._sync_delete_terms(set(term_ids))
        # Add new entries
        self._sync_add_terms(set(term_ids))

    def _embed_all_terms(self) -> None:
        """Embed all terms from self._terms and upsert into DB."""
        self._embed_and_upsert_terms(self._terms)

    def _validate_ids(self) -> None:
        """Validate that all term IDs are unique."""
        ids = []
        for i, term in enumerate(self._terms):
            term_id = term.get("id")
            if term_id is None:
                raise ValueError(f"Term at index {i} is missing an 'id' field.")
            ids.append(term_id)

        # Check for duplicates
        unique_ids = set(ids)
        if len(ids) != len(unique_ids):
            duplicates = [id for id in ids if ids.count(id) > 1]
            raise ValueError(
                f"Duplicate ID(s) found in glossary: {set(duplicates)}. "
                "All IDs must be unique."
            )

        logger.debug(f"Validated {len(unique_ids)} unique IDs.")

    def _load_terms_from_csv(self, csv_path: str) -> None:
        """Read CSV into self._terms list."""
        self._terms = []
        with open(csv_path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            # Skip header row
            next(reader, None)
            for row in reader:
                if len(row) < 3:
                    logger.warning(f"Row has less than 3 columns: {' | '.join(row)}")
                    continue
                try:
                    term_id = int(row[0].strip())
                except (ValueError, TypeError):
                    raise ValueError(
                        f"Invalid ID value '{row[0].strip()}' in CSV. "
                        "ID must be a valid integer."
                    )
                source, target = row[1].strip(), row[2].strip()
                if source and target:
                    self._terms.append(
                        {"id": term_id, "source": source, "target": target}
                    )
        logger.info(f"Read {len(self._terms)} terms from CSV.")

    def _load_terms_from_xlsx(self, xlsx_path: str) -> None:
        """Read XLSX into self._terms list."""
        self._terms = []

        if not os.path.exists(xlsx_path):
            raise FileNotFoundError(f"XLSX file not found: {xlsx_path}")

        try:
            workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
            sheet = workbook.active

            if sheet is None:
                workbook.close()
                raise ValueError("XLSX file is empty or has no active sheet.")

            # Skip header row by enumerating and skipping index 0
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:  # Skip header row
                    continue
                if not row or len(row) < 3:
                    continue
                try:
                    term_id = int(str(row[0]).strip())
                except (ValueError, TypeError):
                    raise ValueError(
                        f"Invalid ID value '{row[0]}' in XLSX at row {i + 1}. "
                        "ID must be a valid integer."
                    )
                source, target = str(row[1]).strip(), str(row[2]).strip()
                if source and target:
                    self._terms.append(
                        {"id": term_id, "source": source, "target": target}
                    )

        finally:
            if "workbook" in locals():
                workbook.close()

        logger.info(f"Read {len(self._terms)} terms from XLSX.")

    def _load_terms_from_xml(self, xml_path: str) -> None:
        """
        Parse a MultiTerm XML file and populate self._terms with glossary entries.

        Each conceptGrp in the XML becomes one entry in self._terms with:
        - "id":     "<filename_stem>_<concept_id>"
        - "source": all terms in the first language joined by " | "
        - "target": all terms in the second language joined by " | "

        Args:
            xml_path: Absolute or relative path to the .xml glossary file.

        Raises:
            FileNotFoundError: If the file does not exist.
            ValueError:        If the XML structure is invalid or inconsistent.
            RuntimeError:      For unexpected errors during parsing.
        """
        # ------------------------------------------------------------------ #
        # 1. Resolve path and derive the file-name stem used as ID prefix     #
        # ------------------------------------------------------------------ #
        path = os.path.abspath(xml_path)
        file_stem = os.path.splitext(os.path.basename(path))[0]
        logger.info("Loading glossary terms from XML: %s (stem=%r)", path, file_stem)

        if not os.path.isfile(path):
            logger.error("XML file not found: %s", path)
            raise FileNotFoundError(f"Glossary XML file not found: {path!r}")

        # ------------------------------------------------------------------ #
        # 2. Parse the XML document                                           #
        # ------------------------------------------------------------------ #
        try:
            tree = ET.parse(path)
        except ET.ParseError as exc:
            logger.error("Malformed XML in %s: %s", path, exc)
            raise ValueError(f"Failed to parse XML file {path!r}: {exc}") from exc
        except OSError as exc:
            logger.error("Cannot read file %s: %s", path, exc)
            raise RuntimeError(f"Cannot read glossary file {path!r}: {exc}") from exc

        root = tree.getroot()
        if root.tag != "mtf":
            logger.error(
                "Unexpected root tag %r in %s (expected 'mtf')", root.tag, path
            )
            raise ValueError(
                f"Invalid glossary XML: root tag is {root.tag!r}, expected 'mtf'."
            )

        # ------------------------------------------------------------------ #
        # 3. Detect the canonical source / target language pair from the      #
        #    first conceptGrp that has at least two distinct language codes.  #
        # ------------------------------------------------------------------ #
        source_lang: str | None = None  # e.g. "EN"
        target_lang: str | None = None  # e.g. "RU"

        concept_groups = root.findall("conceptGrp")
        if not concept_groups:
            logger.warning(
                "No <conceptGrp> elements found in %s; self._terms unchanged.", path
            )
            return

        for cg in concept_groups:
            langs_seen: list[str] = []
            for lg in cg.findall("languageGrp"):
                lang_el = lg.find("language")
                if lang_el is None:
                    continue
                code = lang_el.get("lang", "").strip().upper()
                if code and code not in langs_seen:
                    langs_seen.append(code)
                if len(langs_seen) >= 2:
                    break
            if len(langs_seen) >= 2:
                source_lang, target_lang = langs_seen[0], langs_seen[1]
                logger.debug(
                    "Detected language pair from first qualifying conceptGrp: "
                    "source=%r, target=%r",
                    source_lang,
                    target_lang,
                )
                break

        if source_lang is None or target_lang is None:
            logger.error("Could not detect a source/target language pair in %s", path)
            raise ValueError(
                f"Glossary XML {path!r} does not contain at least two distinct "
                "language codes in any <conceptGrp>."
            )

        # ------------------------------------------------------------------ #
        # 4. Iterate over every <conceptGrp> and build term dicts             #
        # ------------------------------------------------------------------ #
        loaded: list[dict] = []

        for idx, concept_grp in enumerate(concept_groups):
            # --- 4a. Extract the concept ID -------------------------------- #
            concept_el = concept_grp.find("concept")
            if concept_el is None or not (concept_el.text or "").strip():
                logger.warning(
                    "conceptGrp #%d in %s has no <concept> id — skipping.", idx, path
                )
                continue

            assert concept_el.text is not None
            raw_id = concept_el.text.strip()
            entry_id = f"{file_stem}_{raw_id}"

            # --- 4b. Collect terms per language code ----------------------- #
            terms_by_lang: dict[str, list[str]] = {}

            for lg in concept_grp.findall("languageGrp"):
                lang_el = lg.find("language")
                if lang_el is None:
                    logger.debug(
                        "conceptGrp %r: <languageGrp> missing <language> tag — skipped.",
                        entry_id,
                    )
                    continue

                code = lang_el.get("lang", "").strip().upper()
                if not code:
                    logger.debug(
                        "conceptGrp %r: <language> element has no 'lang' attribute — skipped.",
                        entry_id,
                    )
                    continue

                # Warn about unexpected language codes but keep the data
                if code not in (source_lang, target_lang):
                    logger.warning(
                        "conceptGrp %r: unexpected language code %r "
                        "(expected %r or %r) — ignoring.",
                        entry_id,
                        code,
                        source_lang,
                        target_lang,
                    )
                    continue

                term_grp = lg.find("termGrp")
                if term_grp is None:
                    logger.debug(
                        "conceptGrp %r, lang %r: missing <termGrp> — skipped.",
                        entry_id,
                        code,
                    )
                    continue

                term_el = term_grp.find("term")
                if term_el is None or not (term_el.text or "").strip():
                    logger.debug(
                        "conceptGrp %r, lang %r: empty or missing <term> — skipped.",
                        entry_id,
                        code,
                    )
                    continue

                assert term_el.text is not None
                term_text = term_el.text.strip()
                terms_by_lang.setdefault(code, []).append(term_text)

            # --- 4c. Validate that both languages are present -------------- #
            if source_lang not in terms_by_lang:
                logger.warning(
                    "conceptGrp %r: no terms found for source language %r — skipping entry.",
                    entry_id,
                    source_lang,
                )
                continue

            if target_lang not in terms_by_lang:
                logger.warning(
                    "conceptGrp %r: no terms found for target language %r — skipping entry.",
                    entry_id,
                    target_lang,
                )
                continue

            # --- 4d. Join multiple terms with the agreed delimiter --------- #
            source_text = " | ".join(terms_by_lang[source_lang])
            target_text = " | ".join(terms_by_lang[target_lang])

            entry = {"id": entry_id, "source": source_text, "target": target_text}
            loaded.append(entry)
            logger.debug("Parsed entry: %s", entry)

        # ------------------------------------------------------------------ #
        # 5. Commit results                                                   #
        # ------------------------------------------------------------------ #
        self._terms = loaded
        logger.info(
            "Finished loading %d term(s) from %s (source=%r, target=%r).",
            len(loaded),
            path,
            source_lang,
            target_lang,
        )

    def _load_terms_from_file(self, file_path: str) -> None:
        """Auto-detect file format and load accordingly."""
        ext = os.path.splitext(file_path)[1].lower()

        if ext == ".csv":
            self._load_terms_from_csv(file_path)
        elif ext == ".xlsx":
            self._load_terms_from_xlsx(file_path)
        elif ext == ".xml":
            self._load_terms_from_xml(file_path)
        else:
            raise ValueError(
                f"Unsupported file format '{ext}'. "
                "Only .csv and .xlsx files are supported."
            )

    # ── RAG retrieval ─────────────────────────────────────────────────────────

    def retrieve(
        self,
        query_text: str,
        source_lang: str,
        target_lang: str,
        top_k: int = 20,
    ) -> list[dict]:
        """
        Given a text chunk (query), returns the top_k most relevant glossary
        terms as a list of {"source": ..., "target": ...} dicts.

        The returned dicts always have source=source_lang term, target=target_lang term
        regardless of which direction the DB entry was stored in.
        """
        model = self._get_model()
        collection = self._get_collection()

        if collection.count() == 0:
            return []

        # e5 requires "query: " prefix for the search query
        query_embedding = model.encode(
            f"query: {query_text}", normalize_embeddings=True
        ).tolist()

        results = collection.query(
            query_embeddings=[query_embedding],
            n_results=min(
                top_k * 2, collection.count()
            ),  # over-fetch, then deduplicate
            include=["metadatas", "distances"],
        )

        # Check if we got any results
        if not results or not results["metadatas"] or not results["metadatas"][0]:
            return []

        seen_pairs = set()
        terms = []

        for meta in results["metadatas"][0]:
            # Use canonical pair for deduplication (sorted ensures fwd and rev are grouped)
            canonical_pair = (meta["canonical_source"], meta["canonical_target"])
            if canonical_pair in seen_pairs:
                continue
            seen_pairs.add(canonical_pair)

            # Normalize direction: always return {source: src_lang, target: tgt_lang}
            # Compare the requested languages with the canonical pair
            canonical_src, canonical_tgt = canonical_pair

            # Safely convert to strings for comparison
            src_lang_str = str(source_lang).lower()
            tgt_lang_str = str(target_lang).lower()
            canonical_src_str = str(canonical_src).lower()
            canonical_tgt_str = str(canonical_tgt).lower()

            if canonical_src_str == src_lang_str and canonical_tgt_str == tgt_lang_str:
                # Entry already in correct direction
                terms.append({"source": meta["source"], "target": meta["target"]})
            elif (
                canonical_src_str == tgt_lang_str and canonical_tgt_str == src_lang_str
            ):
                # Entry is reversed - swap it
                terms.append({"source": meta["target"], "target": meta["source"]})
            else:
                # Shouldn't happen if canonical pairs are consistent, but return as-is
                terms.append({"source": meta["source"], "target": meta["target"]})

            if len(terms) >= top_k:
                break

        # Log retrieval results
        logger.info(f"Retrieved {len(terms)} glossary terms for chunk")
        if terms:
            pairs_log = "\n".join(f"  - {t['source']} → {t['target']}" for t in terms)
            logger.debug(f"Glossary pairs:\n{pairs_log}")

        return terms

    # ── Utility ───────────────────────────────────────────────────────────────

    def count(self) -> int:
        return self._get_collection().count()

    def clear(self) -> None:
        """Wipe the ChromaDB collection (use before force_reload)."""
        client = chromadb.PersistentClient(
            path=self.chroma_dir,
            settings=Settings(anonymized_telemetry=False),
        )
        client.delete_collection(settings.glossary.collection_name)
        self.collection = None
        logger.info("Collection cleared.")


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Glossary Manager")
    subparsers = parser.add_subparsers(dest="command", help="Available commands")

    # Sync subcommand
    sync_parser = subparsers.add_parser("sync", help="Sync glossary with ChromaDB")
    sync_parser.add_argument(
        "--glossary",
        default="glossary.csv",
        help="Path to glossary CSV or XLSX file",
    )

    args = parser.parse_args()

    if args.command == "sync":
        gm = GlossaryManager()
        gm.sync_glossary(args.glossary, sync_mode=True)
    else:
        parser.print_help()
